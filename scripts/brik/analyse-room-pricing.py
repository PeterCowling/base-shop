#!/usr/bin/env python3
"""
Brikette Room Pricing Analysis
==============================
Source: [OCT-CAL] — Octorate Calendar Export (Price sheet)
Config: [OCT-CFG] — scripts/brik/room-config.json (static room config)

Reads the latest Octorate calendar export and produces a seasonal pricing
summary per room, written to docs/business-os/strategy/BRIK/room-pricing-analysis.md.

Run: python3 scripts/brik/analyse-room-pricing.py
  or: pnpm brik:analyse-room-pricing

The output file is dated and stamped with the source data file used.
Re-running after a fresh Octorate export will overwrite the output with
up-to-date prices.
"""

import glob
import json
import os
import sys
from datetime import date, datetime
from statistics import median

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OCTORATE_DIR = os.path.join(REPO_ROOT, "data", "octorate")
ROOM_INVENTORY_PATH = os.path.join(OCTORATE_DIR, "room-inventory.json")
ROOM_CONFIG_PATH = os.path.join(REPO_ROOT, "scripts", "brik", "room-config.json")
OUTPUT_PATH = os.path.join(REPO_ROOT, "docs", "business-os", "strategy", "BRIK", "room-pricing-analysis.md")

# Seasonal month bands
SEASON_BANDS = {
    "low":  [11, 12, 1, 2, 3],       # Nov-Mar
    "mid":  [4, 5, 9, 10],           # Apr-May, Sep-Oct
    "peak": [6, 7, 8],               # Jun-Aug
}


def find_latest_calendar() -> str:
    pattern = os.path.join(OCTORATE_DIR, "octorate-calendar-*.xlsx")
    files = sorted(glob.glob(pattern), reverse=True)
    if not files:
        print(f"ERROR: No octorate-calendar-*.xlsx found in {OCTORATE_DIR}", file=sys.stderr)
        sys.exit(1)
    return files[0]


def load_room_inventory(path: str) -> dict:
    """Returns {octorate_id: display_label}."""
    with open(path) as f:
        data = json.load(f)
    return {r["id"]: r["name"] for r in data["rooms"]}


def load_room_config(path: str) -> dict:
    """Returns {octorate_id: config_dict}."""
    with open(path) as f:
        data = json.load(f)
    return {r["octorateId"]: r for r in data["rooms"]}


def season_for_month(month: int) -> str:
    for season, months in SEASON_BANDS.items():
        if month in months:
            return season
    return "unknown"


def extract_prices(xlsx_path: str, id_to_config: dict) -> dict:
    """
    Returns {octorate_id: {season: [prices]}} from the Price sheet.
    Rows that have no price (None or 0) are skipped.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Price" not in wb.sheetnames:
        print("ERROR: 'Price' sheet not found in workbook.", file=sys.stderr)
        sys.exit(1)

    ws = wb["Price"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("ERROR: Price sheet is empty.", file=sys.stderr)
        sys.exit(1)

    # Header row: first cell is label/date, remaining cells are room IDs
    header = rows[0]
    room_ids = [str(cell) if cell is not None else None for cell in header[1:]]

    prices_by_room = {rid: {"low": [], "mid": [], "peak": []} for rid in room_ids if rid}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue

        cell_val = row[0]
        if isinstance(cell_val, (date, datetime)):
            row_date = cell_val if isinstance(cell_val, date) else cell_val.date()
        elif isinstance(cell_val, str):
            try:
                row_date = datetime.strptime(cell_val, "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            continue

        month = row_date.month
        season = season_for_month(month)
        if season == "unknown":
            continue

        for i, room_id in enumerate(room_ids):
            if not room_id:
                continue
            raw = row[i + 1]
            if raw is None:
                continue
            try:
                price = float(raw)
            except (TypeError, ValueError):
                continue
            if price <= 0:
                continue
            prices_by_room[room_id][season].append(price)

    wb.close()
    return prices_by_room


def compute_summary(prices_by_room: dict, id_to_config: dict) -> list:
    rows = []
    for room_id, seasons in prices_by_room.items():
        cfg = id_to_config.get(room_id)
        if not cfg:
            continue

        # Skip non-sellable rooms
        if not cfg.get("sellable", True):
            continue

        low_vals  = seasons.get("low", [])
        mid_vals  = seasons.get("mid", [])
        peak_vals = seasons.get("peak", [])

        def fmt(vals):
            if not vals:
                return "—"
            lo = int(min(vals))
            hi = int(max(vals))
            med = int(median(vals))
            if lo == hi:
                return f"EUR {med}"
            return f"EUR {lo}–{hi} (median EUR {med})"

        rows.append({
            "display_name": cfg["displayName"],
            "type":         cfg["type"],
            "gender":       cfg["genderPolicy"],
            "capacity":     cfg["capacity"],
            "low":          fmt(low_vals),
            "mid":          fmt(mid_vals),
            "peak":         fmt(peak_vals),
            "low_n":        len(low_vals),
            "mid_n":        len(mid_vals),
            "peak_n":       len(peak_vals),
            "notes":        cfg.get("notes", ""),
        })

    # Sort: dorms first by capacity desc, then private, then apartment
    order = {"dorm": 0, "private": 1, "apartment": 2}
    rows.sort(key=lambda r: (order.get(r["type"], 9), -(r["capacity"] or 0)))
    return rows


def render_markdown(rows: list, source_file: str, run_date: str) -> str:
    source_basename = os.path.basename(source_file)
    data_date = source_basename.replace("octorate-calendar-", "").replace(".xlsx", "")

    lines = [
        "---",
        "Type: Analysis-Output",
        "Status: Active",
        "Business: BRIK",
        f"Generated: {run_date}",
        f"Source-File: data/octorate/{source_basename}",
        "Source-Data-Date: " + data_date,
        "Script: scripts/brik/analyse-room-pricing.py",
        "---",
        "",
        "# Brikette — Room Pricing Analysis",
        "",
        f"> **Generated:** {run_date}  ",
        f"> **Source data:** `data/octorate/{source_basename}` (Octorate export dated {data_date})  ",
        "> **Script:** `scripts/brik/analyse-room-pricing.py`  ",
        "> **Refresh:** Run `pnpm brik:analyse-room-pricing` after each new Octorate export.",
        "",
        "Prices below are drawn from the `Price` sheet of the Octorate calendar export.",
        "Each cell shows the observed price range across all dates in the season band, plus the median.",
        "These are the headline OTA rates (refundable rate plan). Direct/non-refundable rates may differ.",
        "",
        "**Season bands:**",
        "- Low: Nov–Mar",
        "- Mid: Apr–May, Sep–Oct",
        "- Peak: Jun–Aug",
        "",
        "---",
        "",
        "## Pricing by Room",
        "",
        "| Room | Type | Gender | Cap | Low (Nov–Mar) | Mid (Apr–May, Sep–Oct) | Peak (Jun–Aug) |",
        "|---|---|---|---|---|---|---|",
    ]

    for r in rows:
        cap = str(r["capacity"]) if r["capacity"] is not None else "—"
        lines.append(
            f"| {r['display_name']} | {r['type']} | {r['gender']} | {cap}"
            f" | {r['low']} | {r['mid']} | {r['peak']} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Data Quality Notes",
        "",
        "- Prices are per the OTA refundable rate plan only. Non-refundable and direct-booking rates are not captured in this export.",
        "- Room 8 is excluded (staff accommodation, not for sale in 2026).",
        "- The apartment (2025-14) will appear with no data until it is configured in the Octorate calendar.",
        "- Sample size (n) per season reflects number of date rows with a non-zero price for that room.",
        "  A room showing `—` for a season either has no dates in that season in the export window, or prices are not yet set.",
        "- **Per-bed vs per-room ambiguity:** Smaller rooms price higher than larger rooms (e.g. Room 9 at 3 beds > Room 3 at 8 beds),",
        "  which is consistent with per-bed pricing. Confirm with Pete whether Octorate rates are per bed or per room total.",
        "  This affects ADR and RevPAR calculations.",
        "",
        "---",
        "",
        "## Sample Sizes",
        "",
        "| Room | Low (n) | Mid (n) | Peak (n) |",
        "|---|---|---|---|",
    ]

    for r in rows:
        lines.append(f"| {r['display_name']} | {r['low_n']} | {r['mid_n']} | {r['peak_n']} |")

    lines += [
        "",
        "---",
        "",
        "_This file is auto-generated. Do not edit manually — re-run the script instead._",
        "",
    ]

    return "\n".join(lines)


def main():
    xlsx_path = find_latest_calendar()
    print(f"[brik:analyse-room-pricing] Source: {os.path.relpath(xlsx_path, REPO_ROOT)}")

    inventory = load_room_inventory(ROOM_INVENTORY_PATH)
    id_to_config = load_room_config(ROOM_CONFIG_PATH)

    prices_by_room = extract_prices(xlsx_path, id_to_config)
    rows = compute_summary(prices_by_room, id_to_config)

    run_date = datetime.utcnow().strftime("%Y-%m-%d")
    md = render_markdown(rows, xlsx_path, run_date)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        f.write(md)

    rel_out = os.path.relpath(OUTPUT_PATH, REPO_ROOT)
    print(f"[brik:analyse-room-pricing] Output: {rel_out}")
    print(f"[brik:analyse-room-pricing] Rooms analysed: {len(rows)}")


if __name__ == "__main__":
    main()
